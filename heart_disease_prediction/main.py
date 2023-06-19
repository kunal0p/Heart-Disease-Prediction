from openpyxl import *
from tkinter import *
import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score

wb = load_workbook('D:\\college\\sem_5\\heart_disease\\ex.xlsx')
sheet = wb.active

def excel():

	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50
	sheet.column_dimensions['H'].width = 30
	sheet.column_dimensions['I'].width = 10
	sheet.column_dimensions['J'].width = 10
	sheet.column_dimensions['K'].width = 20
	sheet.column_dimensions['L'].width = 20
	sheet.column_dimensions['M'].width = 40
	sheet.column_dimensions['N'].width = 50

	sheet.cell(row=1, column=1).value = "Age"
	sheet.cell(row=1, column=2).value = "Sex"
	sheet.cell(row=1, column=3).value = "cp"
	sheet.cell(row=1, column=4).value = "trestbps"
	sheet.cell(row=1, column=5).value = "cholestrol"
	sheet.cell(row=1, column=6).value = "fbs"
	sheet.cell(row=1, column=7).value = "restecg"
	sheet.cell(row=1, column=8).value = "thalach"
	sheet.cell(row=1, column=9).value = "exang"
	sheet.cell(row=1, column=10).value = "oldpeak"
	sheet.cell(row=1, column=11).value = "slope"
	sheet.cell(row=1, column=12).value = "ca"
	sheet.cell(row=1, column=13).value = "thal"
	sheet.cell(row=1, column=14).value = "target"



def focus1(event):
	age_field.focus_set()



def focus2(event):
	sex_field.focus_set()



def focus3(event):
	cp_field.focus_set()



def focus4(event):
	trestbps_field.focus_set()

def focus5(event):
	chol_field.focus_set()


def focus6(event):
	fbps_field.focus_set()

def focus7(event):
	restecg_field.focus_set()

def focus8(event):
	thalach_field.focus_set()

def focus9(event):
	exang_field.focus_set()

def focus10(event):
	oldpeak_field.focus_set()

def focus11(event):
	slope_field.focus_set()

def focus12(event):
	ca_field.focus_set()

def focus13(event):
	thal_field.focus_set()


def insert1(a1,a2,a3,a4,a5,a6,a7,a8,a9,a10,a11,a12,a13):
    heart_data=pd.read_csv('D:\college\sem_5\heart_disease\heart.csv')

    x= heart_data.drop(columns='target',axis=1)
    y= heart_data['target']

    x_train,x_test,y_train,y_test = train_test_split(x,y, test_size=0.2, stratify=y, random_state=2)
    model=LogisticRegression()
    model.fit(x_train,y_train)
    
    input_data=(a1,a2,a3,a4,a5,a6,a7,a8,a9,a10,a11,a12,a13)

    #change input data to numpy array
    input_data_array=np.asarray(input_data)

    #reshape the array as we are predicting for only on instance
    input_data_reshape=input_data_array.reshape(1,-1)

    prediction=model.predict(input_data_reshape)
    if(prediction[0]==0):
        target1 = Label(root, text="person have healthy heart", bg="light green")
        target1.grid(row=16, column=4, ipadx="100")
        print("person have healthy heart")
    
    else:
        target1 = Label(root, text="person have unhealthy heart", bg="light green")
        target1.grid(row=16, column=4, ipadx="100")
        print("person have unhealthy heart")

def insert2():
    a1=int(age_field.get())
    a2=int(sex_field.get())
    a3=int(cp_field.get())
    a4=int(trestbps_field.get())
    a5=int(chol_field.get())
    a6=int(fbps_field.get())
    a7=int(restecg_field.get())
    a8=int(thalach_field.get())
    a9=int(exang_field.get())
    a10=float(oldpeak_field.get())
    a11=int(slope_field.get())
    a12=int(ca_field.get())
    a13=int(thal_field.get())
    
    insert1(a1,a2,a3,a4,a5,a6,a7,a8,a9,a10,a11,a12,a13)
def clear():

	age_field.delete(0, END)
	sex_field.delete(0, END)
	cp_field.delete(0, END)
	trestbps_field.delete(0, END)
	chol_field.delete(0, END)
	fbps_field.delete(0, END)
	restecg_field.delete(0, END)
	thalach_field.delete(0, END)
	exang_field.delete(0, END)
	oldpeak_field.delete(0, END)
	trestbps_field.delete(0, END)
	slope_field.delete(0, END)
	ca_field.delete(0, END)
	thal_field.delete(0, END)



def insert():

	if (age_field.get() == "" and
		sex_field.get() == "" and
		cp_field.get() == "" and
		trestbps_field.get() == "" and
		chol_field.get() == "" and
		fbps_field.get() == "" and
		restecg_field.get() == ""and
                thalach_field.get() == "" and
		exang_field.get() == "" and
		oldpeak_field.get() == "" and
		slope_field.get() == "" and
		ca_field.get() == "" and
		thal_field.get() == ""):
			
		print("empty input")

	else:
		current_row = sheet.max_row
		current_column = sheet.max_column

		sheet.cell(row=current_row + 1, column=1).value = age_field.get()
		sheet.cell(row=current_row + 1, column=2).value = sex_field.get()
		sheet.cell(row=current_row + 1, column=3).value = cp_field.get()
		sheet.cell(row=current_row + 1, column=4).value = trestbps_field.get()
		sheet.cell(row=current_row + 1, column=5).value = chol_field.get()
		sheet.cell(row=current_row + 1, column=6).value = fbps_field.get()
		sheet.cell(row=current_row + 1, column=7).value = restecg_field.get()
		sheet.cell(row=current_row + 1, column=8).value = thalach_field.get()
		sheet.cell(row=current_row + 1, column=9).value = exang_field.get()
		sheet.cell(row=current_row + 1, column=10).value = oldpeak_field.get()
		sheet.cell(row=current_row + 1, column=11).value = slope_field.get()
		sheet.cell(row=current_row + 1, column=12).value = ca_field.get()
		sheet.cell(row=current_row + 1, column=13).value = thal_field.get()
		sheet.cell(row=current_row + 1, column=14).value = target_field.get()


		wb.save('D:\\college\\sem_5\\heart_disease\\ex.xlsx')
		age_field.focus_set()
		clear()


# Driver code+
if __name__ == "__main__":
        root = Tk()
        root.configure(background='light green')
        root.title("Heart Disease Prediction")
        root.geometry("500x320")
        excel()

    	# create a Form label
        ding = Label(root, text="Form", bg="light green")

    	# create a Name label
        age = Label(root, text="Age", bg="light green")

        # create a Course label
        sex = Label(root, text="Sex", bg="light green")


    	# create a Form No. label
        cp = Label(root, text="CP", bg="light green")

    	# create a Contact No. label
        trestbps = Label(root, text="trestbps", bg="light green")

    	# create a Email id label
        chol = Label(root, text="chol", bg="light green")

    	# create a address label
        fbps = Label(root, text="fbs", bg="light green")

    	# create a Semester label
        restecg = Label(root, text="restecg", bg="light green")

    	# create a Semester label
        thalach = Label(root, text="thalach", bg="light green")

    	# create a Semester label
        exang = Label(root, text="exang", bg="light green")

    	# create a Semester label
        oldpeak = Label(root, text="oldpeak", bg="light green")

    	# create a Semester label
        slope = Label(root, text="slope", bg="light green")

    	# create a Semester label
        ca= Label(root, text="ca", bg="light green")

    	# create a Semester label
        thal = Label(root, text="thal", bg="light green")

    	# create a Semester label
        target = Label(root, text="target", bg="light green")
        target1 = Label(root, text="result will appear here", bg="light green")
        
        
        
        age.grid(row=3, column=3)
        sex.grid(row=4, column=3)
        cp.grid(row=5, column=3)
        trestbps.grid(row=6, column=3)
        chol.grid(row=7, column=3)
        fbps.grid(row=8, column=3)
        restecg.grid(row=9, column=3)
        thalach.grid(row=10, column=3)
        exang.grid(row=11, column=3)
        oldpeak.grid(row=12, column=3)
        slope.grid(row=13, column=3)
        ca.grid(row=14, column=3)
        thal.grid(row=15, column=3)
        target.grid(row=16, column=3)
        
        
        age_field = Entry(root)
        sex_field = Entry(root)
        cp_field = Entry(root)
        trestbps_field = Entry(root)
        chol_field = Entry(root)
        fbps_field = Entry(root)
        restecg_field = Entry(root)
        exang_field = Entry(root)
        thalach_field = Entry(root)
        oldpeak_field = Entry(root)
        slope_field = Entry(root)
        ca_field = Entry(root)
        thal_field = Entry(root)
        fbps_field = Entry(root)
        target_field = Entry(root)
        
        
        age_field.bind("<Return>", focus1)
        sex_field.bind("<Return>", focus2)
        cp_field.bind("<Return>", focus3)
        trestbps_field.bind("<Return>", focus4)
        chol_field.bind("<Return>", focus5)
        fbps_field.bind("<Return>", focus6)
        restecg_field.bind("<Return>", focus7)
        thalach_field.bind("<Return>", focus8)
        exang_field.bind("<Return>", focus9)
        oldpeak_field.bind("<Return>", focus10)
        slope_field.bind("<Return>", focus11)
        ca_field.bind("<Return>", focus12)
        thal_field.bind("<Return>", focus13)
        
        
        age_field.grid(row=3, column=4, ipadx="100")
        sex_field.grid(row=4, column=4, ipadx="100")
        cp_field.grid(row=5, column=4, ipadx="100")
        trestbps_field.grid(row=6, column=4, ipadx="100")
        chol_field.grid(row=7, column=4, ipadx="100")
        fbps_field.grid(row=8, column=4, ipadx="100")
        restecg_field.grid(row=9, column=4, ipadx="100")
        thalach_field.grid(row=10, column=4, ipadx="100")
        exang_field.grid(row=11, column=4, ipadx="100")
        oldpeak_field.grid(row=12, column=4, ipadx="100")
        slope_field.grid(row=13, column=4, ipadx="100")
        ca_field.grid(row=14, column=4, ipadx="100")
        thal_field.grid(row=15, column=4, ipadx="100")
        target1.grid(row=16, column=4, ipadx="100")
        
        
        excel()
        submit1 = Button(root, text="Submit", fg="Black",
							bg="Red", command=insert2)
        submit = Button(root, text="insert", fg="Black",
							bg="Red", command=insert)
        submit1.grid(row=16, column=5)
        submit.grid(row=16,column=6)
        
        root.mainloop()
